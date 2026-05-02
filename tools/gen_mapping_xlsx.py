"""Generate road_segments_mapping.xlsx for waypoint verification."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROAD_SEGMENTS = [
    # road_id, road_name, segment, origin, destination, distance_meters, via
    ("road_01a", "Shahrae Faisal",                  "Airport to Drigh Road",                    "24.900800,67.168100", "24.887019,67.125427", 5800, None),
    ("road_01b", "Shahrae Faisal",                  "Drigh Road to Karsaz",                     "24.887019,67.125427", "24.874779,67.095790", 3400, None),
    ("road_01c", "Shahrae Faisal",                  "Karsaz to Shaheed-e-Millat Bridge",         "24.874779,67.095790", "24.867330,67.083804", 1500, None),
    ("road_01d", "Shaheed-e-Millat Rd",             "Shaheed-e-Millat Bridge to City School PAF","24.867330,67.083804", "24.861271,67.088200",  850, None),
    ("road_01e", "Shaheed-e-Millat Rd",             "City School PAF to Manzoor Colony Parallel Rd","24.861271,67.088200","24.845614,67.087526",1800, None),
    ("road_01f", "Shahrae Faisal",                  "Shahrah-e-Qaideen Flyover to Metropole",   "24.859755,67.059040", "24.849675,67.030451", 3300, None),
    ("road_02a", "Shaheed-e-Millat Rd",             "Defence/Korangi to Manzoor Colony",        "24.831287,67.079822", "24.852834,67.089831", 2700, None),
    ("road_02b", "Shaheed-e-Millat Rd",             "Manzoor Colony to Shaheed-e-Millat Bridge","24.852834,67.089831", "24.866864,67.083022", 1800, None),
    ("road_04a", "Khayaban-e-Iqbal",               "Metropole to Teen Talwar",                 "24.849675,67.030451", "24.833776,67.033724", 1800, None),
    ("road_04b", "Khayaban-e-Iqbal",               "Teen Talwar to Do Talwar",                 "24.833776,67.033724", "24.821175,67.034203", 1500, None),
    ("road_05a", "Allama Shabbir Ahmed Usmani Rd", "Disco Bakery to Gulshan Chowrangi",        "24.929126,67.097527", "24.924578,67.091605",  800, "24.926800,67.094500"),
    ("road_05b", "Rashid Minhas Road",              "Nipa Chowrangi to Gulshan Chowrangi",      "24.917845,67.097369", "24.924578,67.091605",  900, "24.921200,67.094400"),
    ("road_05c", "Allama Shabbir Ahmed Usmani Rd", "Maskan Chowrangi to Disco Bakery",         "24.935079,67.105263", "24.929126,67.097527", 1000, "24.932000,67.101400"),
    ("road_06a", "Rashid Minhas Road",              "Gulshan Chowrangi to Nipa Chowrangi",      "24.924578,67.091605", "24.917845,67.097369",  900, "24.921200,67.094400"),
    ("road_06b", "Rashid Minhas Road",              "Nipa Chowrangi to Johar Mor",              "24.917845,67.097369", "24.903882,67.114075", 2200, None),
    ("road_06c", "Rashid Minhas Road",              "Johar Mor to Askari 4",                    "24.903882,67.114075", "24.900911,67.116371",  500, "24.902400,67.115200"),
    ("road_06d", "Rashid Minhas Road",              "Askari 4 to Drigh Road",                   "24.900911,67.116371", "24.887019,67.125427", 1500, None),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Road Segments"

# ── Styles ──────────────────────────────────────────────────────────────────
header_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill   = PatternFill("solid", fgColor="1F3864")
via_fill      = PatternFill("solid", fgColor="FFF2CC")   # yellow — needs filling
verified_fill = PatternFill("solid", fgColor="E2EFDA")   # green — pre-populated
no_via_fill   = PatternFill("solid", fgColor="F2F2F2")   # grey — not needed
center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
left          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin          = Side(style="thin", color="CCCCCC")
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Headers ──────────────────────────────────────────────────────────────────
headers = [
    "road_id",
    "Road Name",
    "Segment",
    "Origin (lat,lng)",
    "Destination (lat,lng)",
    "Distance (m)",
    "Via Waypoint Needed?",
    "Via Waypoint (lat,lng)\n[FILL IN VERIFIED COORD]",
    "Notes",
]
col_widths = [12, 28, 42, 22, 22, 14, 20, 34, 40]

for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 36

# ── Data rows ────────────────────────────────────────────────────────────────
for row_idx, (road_id, road_name, segment, origin, destination, distance, via) in enumerate(ROAD_SEGMENTS, start=2):
    via_needed = "Yes" if via else "No"

    row_data = [road_id, road_name, segment, origin, destination, distance, via_needed, via or "", ""]
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = border
        cell.alignment = center if col_idx != 3 else left

    # Colour the Via Waypoint cell
    via_cell = ws.cell(row=row_idx, column=8)
    if via:
        via_cell.fill = verified_fill   # green — pre-populated but unverified
        via_cell.font = Font(name="Calibri", color="375623", italic=True)
    else:
        via_cell.fill = no_via_fill     # grey — not needed

    ws.row_dimensions[row_idx].height = 18

# ── Legend sheet ─────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
legend_rows = [
    ("Colour",           "Meaning"),
    ("Green (col H)",    "Via waypoint pre-populated from current estimate — verify & correct coordinate on Google Maps"),
    ("Grey (col H)",     "No via waypoint needed for this segment"),
    ("Yellow (col H)",   "Reserved — do not use"),
    ("",                 ""),
    ("Instructions",     ""),
    ("1.",               "Open Google Maps in satellite view for each green-highlighted segment"),
    ("2.",               "Find the midpoint of the target road (not a side street)"),
    ("3.",               "Right-click → 'What's here?' to get exact lat,lng"),
    ("4.",               "Paste the verified coordinate into column H (format: lat,lng e.g. 24.926800,67.094500)"),
    ("5.",               "Save the file and share — the pipeline will be rebuilt from these verified coordinates"),
]
ls.column_dimensions["A"].width = 20
ls.column_dimensions["B"].width = 90
for r, (a, b) in enumerate(legend_rows, start=1):
    ls.cell(row=r, column=1, value=a).font = Font(bold=(r == 1 or a.startswith("Instructions")))
    ls.cell(row=r, column=2, value=b)

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/runner/workspace/road_segments_mapping.xlsx"
wb.save(out)
print(f"Saved: {out}")
